"""Convert NSF product CSV/XLSX exports into a normalized JSON dataset.

Usage example:
    python scripts/build_nsf_model_dataset.py \
        --input-csv ./data/nsf_58_export.csv ./data/nsf_53_export_1.xlsx \
        --standard NSF-58 \
        --retrieved-by adama
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _normalize_header(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", name.lower())


def _split_multiline_cell(value: str) -> list[str]:
    parts = re.split(r"\r?\n|;", value)
    cleaned = [part.strip() for part in parts if part.strip()]

    # De-duplicate while preserving order.
    seen: set[str] = set()
    result: list[str] = []
    for item in cleaned:
        key = item.lower()
        if key not in seen:
            seen.add(key)
            result.append(item)
    return result


def _parse_gpd(value: str) -> float | None:
    text = value.strip()
    if not text:
        return None

    match = re.search(r"[0-9]+(?:\.[0-9]+)?", text.replace(",", ""))
    if not match:
        return None

    return float(match.group(0))


def _standard_to_pfas_certified(standard: str) -> bool:
    return standard.upper() in {"NSF-53", "NSF-58"}


def _read_rows(input_csv: Path) -> list[dict[str, str]]:
    if input_csv.suffix.lower() == ".xlsx":
        return _read_xlsx_rows(input_csv)

    with input_csv.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = getattr(dialect, "delimiter", ",")
        if not isinstance(delimiter, str) or len(delimiter) != 1:
            raise csv.Error("invalid delimiter detected")
    except csv.Error:
        dialect = csv.excel

    with input_csv.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, dialect=dialect)
        return [row for row in reader if row]


def _read_xlsx_rows(input_xlsx: Path) -> list[dict[str, str]]:
    workbook = load_workbook(filename=input_xlsx, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)

        headers_raw = next(rows_iter, None)
        if not headers_raw:
            return []

        headers = [str(cell).strip() if cell is not None else "" for cell in headers_raw]
        rows: list[dict[str, str]] = []

        for row_values in rows_iter:
            values = ["" if value is None else str(value).strip() for value in row_values]
            if not any(values):
                continue

            row_dict = {
                headers[i]: values[i] if i < len(values) else ""
                for i in range(len(headers))
                if headers[i]
            }
            if row_dict:
                rows.append(row_dict)

        return rows
    finally:
        workbook.close()


def _value_from_candidates(row: dict[str, str], candidates: list[str]) -> str:
    normalized_map = {_normalize_header(k): v for k, v in row.items() if k}
    for candidate in candidates:
        value = normalized_map.get(candidate)
        if value is not None:
            return value.strip()
    return ""


def _build_record(
    row: dict[str, str],
    standard: str,
    source_document: str,
    source_url: str | None,
) -> dict[str, Any] | None:
    model = _value_from_candidates(
        row,
        [
            "brandnametradenamemodel",
            "model",
            "product",
            "brandmodel",
            "tradenamemodel",
        ],
    )
    if not model:
        return None

    # Some exports repeat table headers inside body rows (across pages); skip those.
    header_markers = {
        "brand name / trade name / model",
        "brand name/trade name/model",
        "model",
    }
    if model.strip().lower() in header_markers:
        return None

    company = _value_from_candidates(row, ["company", "manufacturer", "brand"])
    if company.strip().lower() in {"company", "manufacturer", "brand"}:
        return None

    product_type = _value_from_candidates(row, ["producttype", "type"])
    replacement_module_raw = _value_from_candidates(
        row,
        ["replacementmodule", "replacementmodules"],
    )
    claims_raw = _value_from_candidates(row, ["claims", "claim", "claimsreduction"])
    gpd_raw = _value_from_candidates(row, ["dailyproductionrategpd", "dailyproductionrate"])

    replacements = _split_multiline_cell(replacement_module_raw)
    claims = _split_multiline_cell(claims_raw)

    source_type = "xlsx" if source_document.lower().endswith(".xlsx") else "csv"

    record: dict[str, Any] = {
        "brand": company.upper() if company else "UNKNOWN",
        "model": model,
        "standard": standard,
        "pfas_certified": _standard_to_pfas_certified(standard),
        "product_type": product_type,
        "replacement_modules": replacements,
        "claims": claims,
        "source_type": source_type,
        "source_document": source_document,
    }

    parsed_gpd = _parse_gpd(gpd_raw)
    if parsed_gpd is not None:
        record["daily_production_rate_gpd"] = parsed_gpd

    if source_url:
        record["source_url"] = source_url

    return record


def _dedupe_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[str, str, str]] = set()
    deduped: list[dict[str, Any]] = []

    for record in records:
        key = (
            str(record.get("brand", "")).strip().lower(),
            str(record.get("model", "")).strip().lower(),
            str(record.get("standard", "")).strip().upper(),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(record)

    return deduped


def _load_existing_records(output_json: Path) -> list[dict[str, Any]]:
    if not output_json.exists():
        return []

    try:
        with output_json.open("r", encoding="utf-8") as handle:
            existing = json.load(handle)
    except json.JSONDecodeError:
        return []

    records = existing.get("records", []) if isinstance(existing, dict) else []
    return records if isinstance(records, list) else []


def build_dataset(
    input_files: list[Path],
    output_json: Path,
    standard: str,
    retrieved_by: str,
    source_url: str | None,
) -> dict[str, Any]:
    records: list[dict[str, Any]] = _load_existing_records(output_json)

    for input_file in input_files:
        rows = _read_rows(input_file)
        source_document = input_file.name

        for row in rows:
            record = _build_record(
                row=row,
                standard=standard,
                source_document=source_document,
                source_url=source_url,
            )
            if record is not None:
                records.append(record)

    deduped_records = _dedupe_records(records)

    dataset = {
        "metadata": {
            "name": "NSF Certified Filter Models",
            "generated_at": datetime.now(UTC).isoformat(),
            "retrieved_by": retrieved_by,
            "input_files": [str(path) for path in input_files],
            "standard_filter": standard,
            "source_url": source_url,
            "record_count": len(deduped_records),
        },
        "records": deduped_records,
    }

    output_json.parent.mkdir(parents=True, exist_ok=True)
    with output_json.open("w", encoding="utf-8") as handle:
        json.dump(dataset, handle, indent=2)
        handle.write("\n")

    return dataset


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input-csv",
        required=True,
        nargs="+",
        help="One or more paths to NSF CSV/XLSX exports",
    )
    parser.add_argument(
        "--output-json",
        default="app/data/nsf_certified_models.json",
        help="Output JSON path (default: app/data/nsf_certified_models.json)",
    )
    parser.add_argument(
        "--standard",
        default="NSF-58",
        choices=["NSF-42", "NSF-53", "NSF-58"],
        help="Certification standard represented by the input file",
    )
    parser.add_argument(
        "--retrieved-by",
        required=True,
        help="Name/identifier of dataset retriever",
    )
    parser.add_argument("--source-url", default=None, help="Official source URL for traceability")
    return parser


def main() -> int:
    parser = _build_parser()
    args = parser.parse_args()

    input_files = [Path(item).resolve() for item in args.input_csv]
    output_json = Path(args.output_json).resolve()

    missing = [path for path in input_files if not path.exists()]
    if missing:
        missing_text = ", ".join(str(path) for path in missing)
        parser.error(f"Input file(s) not found: {missing_text}")

    dataset = build_dataset(
        input_files=input_files,
        output_json=output_json,
        standard=args.standard,
        retrieved_by=args.retrieved_by,
        source_url=args.source_url,
    )

    print(f"Wrote {dataset['metadata']['record_count']} records to {output_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
